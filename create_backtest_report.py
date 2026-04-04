"""
create_backtest_report.py — APEX Capital AI Backtest Excel Report
=================================================================
Reads all backtest_*.csv files from logs/ and creates a single
formatted Excel workbook:

  Sheet 1: Summary       — all agents side-by-side, color-coded KPIs
  Sheet 2: GOLD          — full trade log
  Sheet 3: EURUSD        — full trade log
  Sheet 4: GBPUSD        — full trade log
  Sheet 5: USDJPY        — full trade log
  Sheet 6: Monthly P&L   — P&L breakdown by month per agent

Usage:
    python create_backtest_report.py
Output:
    logs/APEX_Backtest_Report.xlsx
"""

import os
import glob
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Config ────────────────────────────────────────────────────────────────────

START_BALANCE = 12200.0
RISK_PCT      = 0.01
RR            = 2.5        # fixed by backtest tp_mult

AGENT_ORDER   = ["GOLD", "EURUSD", "GBPUSD", "USDJPY"]

AGENT_COLORS = {
    "GOLD":   "EAB308",
    "EURUSD": "3B82F6",
    "GBPUSD": "22C55E",
    "USDJPY": "A855F7",
}

# ── Color palette ─────────────────────────────────────────────────────────────
C_BG_DARK     = "0F172A"   # dark navy — header bg
C_BG_MID      = "1E293B"   # mid dark — section header
C_BG_LIGHT    = "1E3A5F"   # blue-dark — subheader
C_BG_ROW_A    = "F8FAFC"   # light grey row
C_BG_ROW_B    = "FFFFFF"   # white row
C_TEXT_LIGHT  = "FFFFFF"   # white text
C_TEXT_DARK   = "0F172A"   # dark text
C_GREEN       = "16A34A"   # positive value
C_RED         = "DC2626"   # negative value
C_AMBER       = "D97706"   # warning/neutral
C_BORDER      = "CBD5E1"   # cell border

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=C_TEXT_DARK, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center")

def border_thin():
    s = Side(border_style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(border_style="medium", color="94A3B8")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Stats computation ─────────────────────────────────────────────────────────

def compute_stats(df: pd.DataFrame, agent_name: str) -> dict:
    total     = len(df)
    wins      = (df['outcome'] == 'WIN').sum()
    losses    = (df['outcome'] == 'LOSS').sum()
    timeouts  = (df['outcome'] == 'TIMEOUT').sum()
    decided   = wins + losses
    win_rate  = wins / decided * 100 if decided > 0 else 0
    ev        = (win_rate / 100 * RR) - ((1 - win_rate / 100) * 1.0)

    # Max consecutive losses
    max_cl = consec = 0
    for outcome in df['outcome']:
        if outcome == 'LOSS':
            consec += 1
            max_cl = max(max_cl, consec)
        elif outcome == 'WIN':
            consec = 0

    # P&L simulation
    risk     = START_BALANCE * RISK_PCT
    balance  = START_BALANCE
    peak     = START_BALANCE
    max_dd   = 0.0
    curve    = [START_BALANCE]
    for outcome in df['outcome']:
        if outcome == 'WIN':
            balance += risk * RR
        elif outcome == 'LOSS':
            balance -= risk
        if balance > peak:
            peak = balance
        dd = (peak - balance) / peak * 100
        if dd > max_dd:
            max_dd = dd
        curve.append(round(balance, 2))

    total_pnl = balance - START_BALANCE
    total_pct = total_pnl / START_BALANCE * 100
    gross_win  = wins  * risk * RR
    gross_loss = losses * risk
    pf = gross_win / gross_loss if gross_loss > 0 else float('inf')

    # Avg bars held (wins vs losses)
    avg_bars_win  = df.loc[df['outcome'] == 'WIN',  'bars_held'].mean() if wins  > 0 else 0
    avg_bars_loss = df.loc[df['outcome'] == 'LOSS', 'bars_held'].mean() if losses > 0 else 0
    avg_adx       = df['h4_adx'].mean()

    # Direction split
    longs  = (df['direction'] == 'BUY').sum()
    shorts = (df['direction'] == 'SELL').sum()

    return {
        "agent":          agent_name,
        "total":          total,
        "wins":           int(wins),
        "losses":         int(losses),
        "timeouts":       int(timeouts),
        "decided":        int(decided),
        "win_rate":       round(win_rate, 1),
        "rr":             RR,
        "ev":             round(ev, 3),
        "max_consec_l":   max_cl,
        "start_balance":  START_BALANCE,
        "end_balance":    round(balance, 2),
        "total_pnl":      round(total_pnl, 2),
        "total_pct":      round(total_pct, 2),
        "max_dd":         round(max_dd, 1),
        "profit_factor":  round(pf, 2) if pf != float('inf') else 999.0,
        "gross_win":      round(gross_win, 2),
        "gross_loss":     round(gross_loss, 2),
        "avg_bars_win":   round(avg_bars_win, 1),
        "avg_bars_loss":  round(avg_bars_loss, 1),
        "avg_adx":        round(avg_adx, 1),
        "longs":          int(longs),
        "shorts":         int(shorts),
        "balance_curve":  curve,
    }


def monthly_pnl(df: pd.DataFrame) -> pd.DataFrame:
    """Returns monthly P&L in $ using 1% risk simulation."""
    risk = START_BALANCE * RISK_PCT
    df2  = df.copy()
    df2['signal_time'] = pd.to_datetime(df2['signal_time'])
    df2['month']       = df2['signal_time'].dt.to_period('M')
    df2['pnl_sim']     = df2['outcome'].map({'WIN': risk * RR, 'LOSS': -risk, 'TIMEOUT': 0})
    return df2.groupby('month')['pnl_sim'].sum().reset_index()


# ── Load all CSVs ─────────────────────────────────────────────────────────────

def _fmt_period(from_str, to_str):
    return f"{from_str[:4]}-{from_str[4:6]}-{from_str[6:]} to {to_str[:4]}-{to_str[4:6]}-{to_str[6:]}"


def load_data():
    """Returns (data, stats, period) using the LONGEST date-range file per agent."""
    data   = {}
    stats  = {}
    period = {}
    start  = {}   # track from-date per agent to pick longest range
    csv_files = glob.glob("logs/backtest_*.csv")

    for fpath in sorted(csv_files):
        fname = os.path.basename(fpath)
        parts = fname.replace(".csv", "").split("_")
        if len(parts) < 4:
            continue
        agent = parts[1]
        if agent not in AGENT_ORDER:
            continue
        from_str, to_str = parts[2], parts[3]
        # Pick the file with the earliest start date (longest history)
        if agent in data and from_str >= start.get(agent, "99999999"):
            continue
        df = pd.read_csv(fpath)
        if df.empty:
            continue
        data[agent]   = df
        stats[agent]  = compute_stats(df, agent)
        period[agent] = _fmt_period(from_str, to_str)
        start[agent]  = from_str

    return data, stats, period


def load_all_runs():
    """Returns dict: {agent: [(label, df, stats, period_str), ...]} — all runs per agent."""
    csv_files = glob.glob("logs/backtest_*.csv")
    runs = {a: [] for a in AGENT_ORDER}

    for fpath in sorted(csv_files):
        fname = os.path.basename(fpath)
        parts = fname.replace(".csv", "").split("_")
        if len(parts) < 4:
            continue
        agent = parts[1]
        if agent not in AGENT_ORDER:
            continue
        df = pd.read_csv(fpath)
        if df.empty:
            continue
        from_str, to_str = parts[2], parts[3]
        # Detect actual M15 coverage from trade data
        actual_from = pd.to_datetime(df['signal_time']).min().strftime('%Y-%m-%d') if len(df) > 0 else from_str[:4]+'-'+from_str[4:6]+'-'+from_str[6:]
        actual_to   = pd.to_datetime(df['signal_time']).max().strftime('%Y-%m-%d') if len(df) > 0 else to_str[:4]+'-'+to_str[4:6]+'-'+to_str[6:]
        label = f"From {from_str[:4]}-{from_str[4:6]}-{from_str[6:]}  (entries: {actual_from} → {actual_to})"
        runs[agent].append((label, df, compute_stats(df, agent), _fmt_period(from_str, to_str)))

    return runs


# ── Sheet builders ────────────────────────────────────────────────────────────

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def write_summary_sheet(ws, stats: dict, period: dict):
    """Sheet 1 — Summary: all agents side by side."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"

    agents = [a for a in AGENT_ORDER if a in stats]

    # ── Title row ─────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(1 + len(agents))}1")
    title_cell = ws["A1"]
    title_cell.value     = "APEX Capital AI — Backtest Summary Report"
    title_cell.font      = font(bold=True, color=C_TEXT_LIGHT, size=14)
    title_cell.fill      = fill(C_BG_DARK)
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 28

    # ── Period row ────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(1 + len(agents))}2")
    periods = list({v for v in period.values()})
    period_str = periods[0] if len(periods) == 1 else "Multiple periods"
    period_cell = ws["A2"]
    period_cell.value     = f"Period: {period_str}  |  Balance: ${START_BALANCE:,.0f}  |  Risk: {RISK_PCT*100:.0f}% per trade  |  R:R: {RR}x fixed"
    period_cell.font      = font(color=C_TEXT_LIGHT, size=10, italic=True)
    period_cell.fill      = fill(C_BG_MID)
    period_cell.alignment = center()
    ws.row_dimensions[2].height = 18

    # ── Agent name headers ────────────────────────────────────────
    ws["A3"].value     = "Metric"
    ws["A3"].font      = font(bold=True, color=C_TEXT_LIGHT)
    ws["A3"].fill      = fill(C_BG_LIGHT)
    ws["A3"].alignment = center()
    ws["A3"].border    = border_thin()
    ws.row_dimensions[3].height = 22

    for col_idx, agent in enumerate(agents, start=2):
        cell = ws.cell(row=3, column=col_idx)
        cell.value     = agent
        cell.font      = font(bold=True, color=C_TEXT_DARK, size=12)
        cell.fill      = fill(AGENT_COLORS[agent])
        cell.alignment = center()
        cell.border    = border_thin()

    # ── Sections and metrics ──────────────────────────────────────
    SECTIONS = [
        ("SIGNAL QUALITY", [
            ("Total Signals",       "total",         "int",   None),
            ("Wins",                "wins",          "int",   None),
            ("Losses",              "losses",        "int",   None),
            ("Timeouts (>50h)",     "timeouts",      "int",   None),
            ("Win Rate (excl. TO)", "win_rate",      "pct1",  (40, 200)),   # green if >40
            ("Expected Value",      "ev",            "r2+",   (0, 200)),    # green if >0
            ("Max Consec. Losses",  "max_consec_l",  "int",   (-1, 5)),     # green if <=5
            ("Longs (BUY)",         "longs",         "int",   None),
            ("Shorts (SELL)",       "shorts",        "int",   None),
        ]),
        ("PERFORMANCE", [
            ("Start Balance ($)",   "start_balance", "usd",   None),
            ("End Balance ($)",     "end_balance",   "usd",   None),
            ("Total P&L ($)",       "total_pnl",     "usd+",  (0, 200)),
            ("Total Return (%)",    "total_pct",     "pct1+", (0, 200)),
            ("Max Drawdown (%)",    "max_dd",        "pct1",  (-1, 10)),    # green if <10
            ("Profit Factor",       "profit_factor", "r2",    (1.0, 200)),  # green if >=1
            ("Gross Win ($)",       "gross_win",     "usd",   None),
            ("Gross Loss ($)",      "gross_loss",    "usd",   None),
        ]),
        ("TRADE ANALYSIS", [
            ("Avg Bars Held (Win)",  "avg_bars_win",  "r1",   None),
            ("Avg Bars Held (Loss)", "avg_bars_loss", "r1",   None),
            ("Avg H4 ADX",           "avg_adx",       "r1",   (25, 200)),   # green if >25
        ]),
    ]

    row = 4
    for section_name, metrics in SECTIONS:
        # Section header
        ws.merge_cells(f"A{row}:{get_column_letter(1 + len(agents))}{row}")
        sc = ws.cell(row=row, column=1)
        sc.value     = section_name
        sc.font      = font(bold=True, color=C_TEXT_LIGHT, size=10)
        sc.fill      = fill(C_BG_MID)
        sc.alignment = center()
        ws.row_dimensions[row].height = 18
        row += 1

        for metric_label, key, fmt, threshold in metrics:
            ws.row_dimensions[row].height = 20
            bg = fill(C_BG_ROW_A) if row % 2 == 0 else fill(C_BG_ROW_B)

            # Label cell
            lc = ws.cell(row=row, column=1)
            lc.value     = metric_label
            lc.font      = font(size=10)
            lc.fill      = bg
            lc.alignment = left()
            lc.border    = border_thin()

            for col_idx, agent in enumerate(agents, start=2):
                val  = stats[agent].get(key, "N/A")
                cell = ws.cell(row=row, column=col_idx)
                cell.fill      = bg
                cell.alignment = center()
                cell.border    = border_thin()

                # Format value
                if fmt == "int":
                    cell.value        = int(val) if val != "N/A" else "N/A"
                    cell.number_format = "0"
                elif fmt == "pct1":
                    cell.value        = float(val)
                    cell.number_format = '0.0"%"'
                elif fmt == "pct1+":
                    cell.value        = float(val)
                    cell.number_format = '+0.0"%";-0.0"%"'
                elif fmt == "r2":
                    cell.value        = float(val)
                    cell.number_format = "0.00"
                elif fmt == "r2+":
                    cell.value        = float(val)
                    cell.number_format = '+0.000;-0.000'
                elif fmt == "r1":
                    cell.value        = float(val)
                    cell.number_format = "0.0"
                elif fmt == "usd":
                    cell.value        = float(val)
                    cell.number_format = '"$"#,##0.00'
                elif fmt == "usd+":
                    cell.value        = float(val)
                    cell.number_format = '"$"+#,##0.00;"$"-#,##0.00'
                else:
                    cell.value = val

                # Color coding based on threshold
                if threshold is not None and val != "N/A":
                    lo, hi = threshold
                    v = float(val)
                    if key == "max_consec_l":
                        # lower is better
                        cell.font = font(bold=True, color=C_GREEN if v <= lo else C_RED, size=10)
                    elif key == "max_dd":
                        # lower is better
                        cell.font = font(bold=True, color=C_GREEN if v < lo else C_RED, size=10)
                    else:
                        # higher is better
                        cell.font = font(bold=True, color=C_GREEN if v >= lo else C_RED, size=10)
                else:
                    cell.font = font(size=10)

            row += 1

    # ── Column widths ──────────────────────────────────────────────
    set_col_width(ws, 1, 26)
    for c in range(2, 2 + len(agents)):
        set_col_width(ws, c, 18)

    # ── Verdict row ────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:{get_column_letter(1 + len(agents))}{row}")
    ws.cell(row=row, column=1).value = "VERDICT"
    ws.cell(row=row, column=1).font  = font(bold=True, color=C_TEXT_LIGHT, size=10)
    ws.cell(row=row, column=1).fill  = fill(C_BG_MID)
    ws.cell(row=row, column=1).alignment = center()
    ws.row_dimensions[row].height = 18
    row += 1

    verdicts = {
        "GOLD":   ("✅ DEPLOY",   C_GREEN,  "Positive EV, PF > 1, drawdown manageable"),
        "EURUSD": ("⚠️ TUNE",    C_AMBER,  "Marginally break-even — signal rules need tightening"),
        "GBPUSD": ("⚠️ TUNE",    C_AMBER,  "Slightly negative — high timeout rate (43%), refine entry"),
        "USDJPY": ("❌ PAUSE",    C_RED,    "Very low win rate (5%) — strategy not suited to 2025 USDJPY chop"),
    }

    for col_idx, agent in enumerate(agents, start=2):
        v_label, v_color, v_note = verdicts.get(agent, ("—", C_TEXT_DARK, ""))
        ws.row_dimensions[row].height = 20
        vc = ws.cell(row=row, column=col_idx)
        vc.value     = v_label
        vc.font      = font(bold=True, color=v_color, size=11)
        vc.fill      = fill(C_BG_ROW_A if row % 2 == 0 else C_BG_ROW_B)
        vc.alignment = center()
        vc.border    = border_thin()

    ws.cell(row=row, column=1).value     = "Verdict"
    ws.cell(row=row, column=1).font      = font(bold=True, size=10)
    ws.cell(row=row, column=1).fill      = fill(C_BG_ROW_A if row % 2 == 0 else C_BG_ROW_B)
    ws.cell(row=row, column=1).alignment = left()
    ws.cell(row=row, column=1).border    = border_thin()
    row += 1

    # Verdict notes
    for col_idx, agent in enumerate(agents, start=2):
        _, _, v_note = verdicts.get(agent, ("—", C_TEXT_DARK, ""))
        ws.row_dimensions[row].height = 32
        nc = ws.cell(row=row, column=col_idx)
        nc.value     = v_note
        nc.font      = font(italic=True, size=9, color="475569")
        nc.fill      = fill("F1F5F9")
        nc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        nc.border    = border_thin()

    ws.cell(row=row, column=1).value     = "Notes"
    ws.cell(row=row, column=1).font      = font(size=10, italic=True)
    ws.cell(row=row, column=1).fill      = fill("F1F5F9")
    ws.cell(row=row, column=1).alignment = left()
    ws.cell(row=row, column=1).border    = border_thin()


def write_trade_sheet(ws, df: pd.DataFrame, agent_name: str):
    """One sheet per agent — full trade log."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    color = AGENT_COLORS[agent_name]

    # Title
    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value     = f"{agent_name} — Full Trade Log"
    tc.font      = font(bold=True, color=C_TEXT_LIGHT, size=13)
    tc.fill      = fill(color)
    tc.alignment = center()
    ws.row_dimensions[1].height = 26

    # Headers
    headers = [
        "Signal Time", "Exit Time", "Direction", "Entry Price",
        "SL (pts)", "TP (pts)", "Outcome", "P&L (pts)",
        "Bars Held", "H4 ADX", "H4 RSI", "H4 ATR"
    ]
    col_widths = [18, 18, 10, 13, 10, 10, 10, 12, 10, 9, 9, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value     = h
        cell.font      = font(bold=True, color=C_TEXT_LIGHT, size=10)
        cell.fill      = fill(C_BG_LIGHT)
        cell.alignment = center()
        cell.border    = border_thin()
        set_col_width(ws, col_idx, w)

    ws.row_dimensions[2].height = 18

    # Data rows
    outcome_colors = {"WIN": "DCFCE7", "LOSS": "FEE2E2", "TIMEOUT": "FEF3C7"}
    outcome_fonts  = {"WIN": C_GREEN,  "LOSS": C_RED,    "TIMEOUT": C_AMBER}

    for r_idx, row in df.iterrows():
        excel_row = r_idx + 3
        ws.row_dimensions[excel_row].height = 16
        outcome = row.get('outcome', '')
        row_bg  = outcome_colors.get(outcome, C_BG_ROW_A)

        values = [
            row.get('signal_time', ''),
            row.get('exit_time',   ''),
            row.get('direction',   ''),
            row.get('entry_price', ''),
            row.get('sl_pts',      ''),
            row.get('tp_pts',      ''),
            outcome,
            row.get('pnl_pts',    ''),
            row.get('bars_held',  ''),
            row.get('h4_adx',     ''),
            row.get('h4_rsi',     ''),
            row.get('h4_atr',     ''),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value     = val
            cell.fill      = fill(row_bg)
            cell.border    = border_thin()
            cell.alignment = center()
            if col_idx == 7:  # Outcome column
                cell.font = font(bold=True, color=outcome_fonts.get(outcome, C_TEXT_DARK), size=10)
            else:
                cell.font = font(size=10)


def write_monthly_sheet(ws, data: dict):
    """Monthly P&L breakdown per agent."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    agents = [a for a in AGENT_ORDER if a in data]

    # Title
    ws.merge_cells(f"A1:{get_column_letter(1 + len(agents))}1")
    tc = ws["A1"]
    tc.value     = "Monthly P&L by Agent  ($)"
    tc.font      = font(bold=True, color=C_TEXT_LIGHT, size=13)
    tc.fill      = fill(C_BG_DARK)
    tc.alignment = center()
    ws.row_dimensions[1].height = 26

    # Gather all months across all agents
    monthly_data = {}
    all_months   = set()
    for agent in agents:
        mpnl = monthly_pnl(data[agent])
        monthly_data[agent] = dict(zip(mpnl['month'].astype(str), mpnl['pnl_sim']))
        all_months.update(monthly_data[agent].keys())

    all_months = sorted(all_months)

    # Headers
    ws.cell(row=2, column=1).value     = "Month"
    ws.cell(row=2, column=1).font      = font(bold=True, color=C_TEXT_LIGHT)
    ws.cell(row=2, column=1).fill      = fill(C_BG_LIGHT)
    ws.cell(row=2, column=1).alignment = center()
    ws.cell(row=2, column=1).border    = border_thin()
    set_col_width(ws, 1, 14)

    for col_idx, agent in enumerate(agents, start=2):
        cell = ws.cell(row=2, column=col_idx)
        cell.value     = agent
        cell.font      = font(bold=True, color=C_TEXT_DARK)
        cell.fill      = fill(AGENT_COLORS[agent])
        cell.alignment = center()
        cell.border    = border_thin()
        set_col_width(ws, col_idx, 14)

    # Add TOTAL column
    total_col = 2 + len(agents)
    tc2 = ws.cell(row=2, column=total_col)
    tc2.value     = "TOTAL"
    tc2.font      = font(bold=True, color=C_TEXT_LIGHT)
    tc2.fill      = fill(C_BG_MID)
    tc2.alignment = center()
    tc2.border    = border_thin()
    set_col_width(ws, total_col, 14)
    ws.row_dimensions[2].height = 20

    # Data rows
    row_totals = []
    for r_idx, month in enumerate(all_months):
        excel_row = r_idx + 3
        ws.row_dimensions[excel_row].height = 18

        ws.cell(row=excel_row, column=1).value     = month
        ws.cell(row=excel_row, column=1).font      = font(size=10)
        ws.cell(row=excel_row, column=1).alignment = center()
        ws.cell(row=excel_row, column=1).border    = border_thin()
        ws.cell(row=excel_row, column=1).fill      = fill(C_BG_ROW_A if excel_row % 2 == 0 else C_BG_ROW_B)

        row_total = 0
        for col_idx, agent in enumerate(agents, start=2):
            val  = monthly_data[agent].get(month, 0)
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value        = round(val, 2)
            cell.number_format = '"$"#,##0.00;[Red]"$"-#,##0.00'
            cell.font         = font(bold=False, color=C_GREEN if val >= 0 else C_RED, size=10)
            cell.alignment    = center()
            cell.border       = border_thin()
            cell.fill         = fill(C_BG_ROW_A if excel_row % 2 == 0 else C_BG_ROW_B)
            row_total += val

        # Total column
        tot_cell = ws.cell(row=excel_row, column=total_col)
        tot_cell.value        = round(row_total, 2)
        tot_cell.number_format = '"$"#,##0.00;[Red]"$"-#,##0.00'
        tot_cell.font         = font(bold=True, color=C_GREEN if row_total >= 0 else C_RED, size=10)
        tot_cell.alignment    = center()
        tot_cell.border       = border_thin()
        tot_cell.fill         = fill("EFF6FF" if excel_row % 2 == 0 else "DBEAFE")
        row_totals.append(row_total)

    # Grand total row
    grand_row = 3 + len(all_months)
    ws.row_dimensions[grand_row].height = 22
    ws.merge_cells(f"A{grand_row}:A{grand_row}")
    gc = ws.cell(row=grand_row, column=1)
    gc.value     = "TOTAL"
    gc.font      = font(bold=True, color=C_TEXT_LIGHT, size=11)
    gc.fill      = fill(C_BG_MID)
    gc.alignment = center()
    gc.border    = border_medium()

    for col_idx, agent in enumerate(agents, start=2):
        agent_total = sum(monthly_data[agent].get(m, 0) for m in all_months)
        cell = ws.cell(row=grand_row, column=col_idx)
        cell.value        = round(agent_total, 2)
        cell.number_format = '"$"#,##0.00;[Red]"$"-#,##0.00'
        cell.font         = font(bold=True, color=C_GREEN if agent_total >= 0 else C_RED, size=11)
        cell.alignment    = center()
        cell.border       = border_medium()
        cell.fill         = fill(C_BG_MID)

    grand_total = sum(row_totals)
    gtc = ws.cell(row=grand_row, column=total_col)
    gtc.value        = round(grand_total, 2)
    gtc.number_format = '"$"#,##0.00;[Red]"$"-#,##0.00'
    gtc.font         = font(bold=True, color=C_GREEN if grand_total >= 0 else C_RED, size=12)
    gtc.alignment    = center()
    gtc.border       = border_medium()
    gtc.fill         = fill(C_BG_DARK)


def write_comparison_sheet(ws, all_runs: dict):
    """
    Sheet: Run Comparison — shows every backtest run per agent side by side.
    Highlights the data-availability limitation for multi-year runs.
    """
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    KEY_METRICS = [
        ("Requested From",     "period_from",    "str"),
        ("Actual M15 From",    "actual_from",    "str"),
        ("Actual M15 To",      "actual_to",      "str"),
        ("Trades",             "total",          "int"),
        ("Win Rate",           "win_rate",       "pct1"),
        ("Expected Value",     "ev",             "r3+"),
        ("Total P&L ($)",      "total_pnl",      "usd+"),
        ("Total Return (%)",   "total_pct",      "pct1+"),
        ("Max Drawdown (%)",   "max_dd",         "pct1"),
        ("Profit Factor",      "profit_factor",  "r2"),
        ("Max Consec. Loss",   "max_consec_l",   "int"),
    ]

    # Title
    total_cols = 1 + sum(len(v) for v in all_runs.values() if v)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    tc = ws["A1"]
    tc.value     = "APEX Capital AI — Backtest Run Comparison"
    tc.font      = font(bold=True, color=C_TEXT_LIGHT, size=13)
    tc.fill      = fill(C_BG_DARK)
    tc.alignment = center()
    ws.row_dimensions[1].height = 26

    # Note row
    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    nc = ws["A2"]
    nc.value     = ("NOTE: MetaQuotes demo stores ~5-15 months of M15 data. "
                    "Requesting a longer 'From' date improves H4/H1 indicator convergence "
                    "but does NOT extend M15 entry coverage beyond what's cached.")
    nc.font      = font(italic=True, size=9, color="475569")
    nc.fill      = fill("FFF7ED")
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    nc.border    = border_thin()
    ws.row_dimensions[2].height = 30

    # Build column layout: agent headers spanning their runs
    col = 2
    agent_col_map = {}  # agent -> list of col indices
    for agent in AGENT_ORDER:
        runs = all_runs.get(agent, [])
        if not runs:
            continue
        agent_col_map[agent] = list(range(col, col + len(runs)))
        # Agent header (spanning all runs for this agent)
        if len(runs) > 1:
            ws.merge_cells(f"{get_column_letter(col)}3:{get_column_letter(col+len(runs)-1)}3")
        ac = ws.cell(row=3, column=col)
        ac.value     = agent
        ac.font      = font(bold=True, color=C_TEXT_DARK, size=12)
        ac.fill      = fill(AGENT_COLORS[agent])
        ac.alignment = center()
        ac.border    = border_thin()
        col += len(runs)

    ws.cell(row=3, column=1).value     = "Metric"
    ws.cell(row=3, column=1).font      = font(bold=True, color=C_TEXT_LIGHT)
    ws.cell(row=3, column=1).fill      = fill(C_BG_LIGHT)
    ws.cell(row=3, column=1).alignment = center()
    ws.cell(row=3, column=1).border    = border_thin()
    ws.row_dimensions[3].height = 22

    # Run sub-headers (row 4)
    col = 2
    for agent in AGENT_ORDER:
        runs = all_runs.get(agent, [])
        for i, (label, df, s, pstr) in enumerate(runs):
            rc = ws.cell(row=4, column=col)
            short = f"Run {i+1}: {pstr.split(' to ')[0].strip()}"
            rc.value     = short
            rc.font      = font(bold=True, color=C_TEXT_LIGHT, size=9)
            rc.fill      = fill(C_BG_MID)
            rc.alignment = center()
            rc.border    = border_thin()
            col += 1

    ws.cell(row=4, column=1).value     = ""
    ws.cell(row=4, column=1).fill      = fill(C_BG_MID)
    ws.cell(row=4, column=1).border    = border_thin()
    ws.row_dimensions[4].height = 18

    # Data rows
    for r_idx, (metric_label, key, fmt) in enumerate(KEY_METRICS):
        excel_row = r_idx + 5
        ws.row_dimensions[excel_row].height = 18
        bg = fill(C_BG_ROW_A if excel_row % 2 == 0 else C_BG_ROW_B)

        lc = ws.cell(row=excel_row, column=1)
        lc.value     = metric_label
        lc.font      = font(size=10)
        lc.fill      = bg
        lc.alignment = left()
        lc.border    = border_thin()

        col = 2
        for agent in AGENT_ORDER:
            runs = all_runs.get(agent, [])
            for i, (label, df, s, pstr) in enumerate(runs):
                cell = ws.cell(row=excel_row, column=col)
                cell.fill      = bg
                cell.alignment = center()
                cell.border    = border_thin()

                # Derive special fields
                if key == "period_from":
                    val = pstr.split(" to ")[0]
                elif key == "actual_from":
                    val = pd.to_datetime(df['signal_time']).min().strftime('%Y-%m-%d') if len(df) > 0 else "N/A"
                elif key == "actual_to":
                    val = pd.to_datetime(df['signal_time']).max().strftime('%Y-%m-%d') if len(df) > 0 else "N/A"
                else:
                    val = s.get(key, "N/A")

                if fmt == "str":
                    cell.value = str(val)
                    cell.font  = font(size=9, italic=True, color="475569")
                elif fmt == "int":
                    cell.value = int(val) if val != "N/A" else "N/A"
                    cell.font  = font(size=10)
                elif fmt == "pct1":
                    cell.value        = float(val)
                    cell.number_format = '0.0"%"'
                    cell.font = font(bold=True, color=C_GREEN if float(val) >= 40 else C_RED, size=10)
                elif fmt == "pct1+":
                    cell.value        = float(val)
                    cell.number_format = '+0.0"%";-0.0"%"'
                    cell.font = font(bold=True, color=C_GREEN if float(val) >= 0 else C_RED, size=10)
                elif fmt == "r3+":
                    cell.value        = float(val)
                    cell.number_format = '+0.000;-0.000'
                    cell.font = font(bold=True, color=C_GREEN if float(val) > 0 else C_RED, size=10)
                elif fmt == "r2":
                    cell.value        = float(val)
                    cell.number_format = "0.00"
                    cell.font = font(bold=True, color=C_GREEN if float(val) >= 1.0 else C_RED, size=10)
                elif fmt == "usd+":
                    cell.value        = float(val)
                    cell.number_format = '"$"+#,##0.00;"$"-#,##0.00'
                    cell.font = font(bold=True, color=C_GREEN if float(val) >= 0 else C_RED, size=10)
                else:
                    cell.value = val
                    cell.font  = font(size=10)
                col += 1

    # Column widths
    set_col_width(ws, 1, 22)
    col = 2
    for agent in AGENT_ORDER:
        for _ in all_runs.get(agent, []):
            set_col_width(ws, col, 20)
            col += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n[REPORT] Loading backtest CSV files...")
    data, stats, period = load_data()
    all_runs = load_all_runs()

    if not data:
        print("[REPORT] ERROR: No backtest CSV files found in logs/")
        print("         Run: python backtest.py --all --from 2025-01-01")
        return

    agents_found = [a for a in AGENT_ORDER if a in data]
    print(f"[REPORT] Found: {', '.join(agents_found)}")

    wb = Workbook()

    # ── Sheet 1: Summary (longest run per agent) ──────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    print("[REPORT] Building Summary sheet...")
    write_summary_sheet(ws_summary, stats, period)

    # ── Sheet 2: Run Comparison ───────────────────────────────────
    ws_compare = wb.create_sheet(title="Run Comparison")
    print("[REPORT] Building Run Comparison sheet...")
    write_comparison_sheet(ws_compare, all_runs)

    # ── Sheets 3-6: Trade logs ────────────────────────────────────
    for agent in agents_found:
        ws = wb.create_sheet(title=agent)
        print(f"[REPORT] Building {agent} trade log sheet...")
        write_trade_sheet(ws, data[agent], agent)

    # ── Sheet 7: Monthly P&L ──────────────────────────────────────
    ws_monthly = wb.create_sheet(title="Monthly P&L")
    print("[REPORT] Building Monthly P&L sheet...")
    write_monthly_sheet(ws_monthly, data)

    # ── Save ──────────────────────────────────────────────────────
    out_path = "logs/APEX_Backtest_Report.xlsx"
    wb.save(out_path)
    print(f"\n[REPORT] DONE. Saved -> {out_path}")
    print(f"[REPORT] Sheets: Summary | Run Comparison | {' | '.join(agents_found)} | Monthly P&L")


if __name__ == "__main__":
    main()
